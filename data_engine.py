import io
import os
import shutil
from datetime import date, datetime, timedelta

import pandas as pd
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PORTFOLIOS_DIR = "portfolios"


def _ensure_dir():
    os.makedirs(PORTFOLIOS_DIR, exist_ok=True)


def _portfolio_path(name: str) -> str:
    safe = "".join(c for c in name if c.isalnum() or c in "-_ ")
    return os.path.join(PORTFOLIOS_DIR, f"{safe}.csv")


def list_portfolio_names() -> list[str]:
    _ensure_dir()
    legacy = "portfolio.csv"
    if os.path.exists(legacy) and not any(f.endswith(".csv") for f in os.listdir(PORTFOLIOS_DIR)):
        shutil.copy(legacy, _portfolio_path("Main"))
    return sorted(f[:-4] for f in os.listdir(PORTFOLIOS_DIR) if f.endswith(".csv"))


def load_portfolio(name: str) -> pd.DataFrame:
    path = _portfolio_path(name)
    if not os.path.exists(path):
        return pd.DataFrame(columns=["ID", "Ticker", "Shares", "Cost Basis", "Purchase Date"])
    df = pd.read_csv(path, dtype=str)
    df.columns = df.columns.str.strip()

    changed = False
    if "ID" not in df.columns:
        df.insert(0, "ID", [str(i) for i in range(1, len(df) + 1)])
        changed = True
    if "Purchase Date" not in df.columns:
        df["Purchase Date"] = ""
        changed = True

    df["Ticker"] = df["Ticker"].str.upper().str.strip()
    df["Shares"] = pd.to_numeric(df["Shares"], errors="coerce")
    df["Cost Basis"] = pd.to_numeric(df["Cost Basis"], errors="coerce")
    df["Purchase Date"] = df["Purchase Date"].fillna("").str.strip()

    if changed:
        save_portfolio(df, name)
    return df


def save_portfolio(df: pd.DataFrame, name: str) -> None:
    _ensure_dir()
    df.to_csv(_portfolio_path(name), index=False)


def _next_id(df: pd.DataFrame) -> str:
    if df.empty or "ID" not in df.columns:
        return "1"
    valid = [int(i) for i in df["ID"].astype(str) if i.isdigit()]
    return str(max(valid, default=0) + 1)


def create_portfolio(name: str) -> None:
    _ensure_dir()
    path = _portfolio_path(name)
    if os.path.exists(path):
        raise ValueError(f"Portfolio '{name}' already exists")
    pd.DataFrame(columns=["ID", "Ticker", "Shares", "Cost Basis", "Purchase Date"]).to_csv(path, index=False)


def delete_portfolio(name: str) -> None:
    path = _portfolio_path(name)
    if os.path.exists(path):
        os.remove(path)


def add_holding(name: str, ticker: str, shares: float, cost_basis: float, purchase_date: str = "") -> None:
    df = load_portfolio(name)
    new_row = pd.DataFrame([{
        "ID": _next_id(df),
        "Ticker": ticker.upper().strip(),
        "Shares": shares,
        "Cost Basis": cost_basis,
        "Purchase Date": (purchase_date or "").strip(),
    }])
    df = pd.concat([df, new_row], ignore_index=True)
    save_portfolio(df, name)


def update_holding(name: str, lot_id: str, shares: float, cost_basis: float, purchase_date: str = "") -> None:
    df = load_portfolio(name)
    mask = df["ID"].astype(str) == str(lot_id)
    if not mask.any():
        raise ValueError(f"Lot ID {lot_id} not found")
    df.loc[mask, "Shares"] = shares
    df.loc[mask, "Cost Basis"] = cost_basis
    df.loc[mask, "Purchase Date"] = (purchase_date or "").strip()
    save_portfolio(df, name)


def remove_holding(name: str, lot_id: str) -> None:
    df = load_portfolio(name)
    df = df[df["ID"].astype(str) != str(lot_id)]
    save_portfolio(df, name)


def _download_close(tickers: list[str], **kwargs) -> pd.DataFrame:
    data = yf.download(tickers, auto_adjust=True, progress=False, **kwargs)
    close = data["Close"]
    if isinstance(close, pd.Series):
        close = close.to_frame(name=tickers[0])
    return close


def fetch_current_prices(tickers: list[str]) -> dict[str, float | None]:
    if not tickers:
        return {}
    close = _download_close(tickers, period="2d")
    return {t: (float(close[t].dropna().iloc[-1]) if t in close else None) for t in tickers}


def fetch_one_year_returns(tickers: list[str]) -> dict[str, float | None]:
    if not tickers:
        return {}
    end = datetime.today()
    start = end - timedelta(days=365)
    close = _download_close(tickers, start=start.strftime("%Y-%m-%d"), end=end.strftime("%Y-%m-%d"))
    returns = {}
    for t in tickers:
        try:
            col = close[t].dropna()
            returns[t] = round((float(col.iloc[-1]) / float(col.iloc[0]) - 1) * 100, 2) if len(col) >= 2 else None
        except (KeyError, IndexError):
            returns[t] = None
    return returns


def fetch_price_history(name: str, period: str = "1y", start: str | None = None, end: str | None = None) -> dict:
    tickers = load_portfolio(name)["Ticker"].unique().tolist()
    if not tickers:
        return {"dates": [], "series": {}}
    if start and end:
        close = _download_close(tickers, start=start, end=end)
    else:
        close = _download_close(tickers, period=period)
    dates = [d.strftime("%Y-%m-%d") for d in close.index]
    series = {}
    for t in tickers:
        try:
            col = close[t]
            first = col.first_valid_index()
            if first is None:
                continue
            base = float(col[first])
            series[t] = [round((float(v) / base - 1) * 100, 2) if pd.notna(v) else None for v in col]
        except KeyError:
            pass
    return {"dates": dates, "series": series}


def _days_held(purchase_date_str) -> int | None:
    s = str(purchase_date_str or "").strip()
    if not s:
        return None
    try:
        return (date.today() - date.fromisoformat(s)).days
    except ValueError:
        return None


def _annualized_return(pnl_pct: float | None, days: int | None) -> float | None:
    if pnl_pct is None or days is None or days <= 0:
        return None
    try:
        return round(((1 + pnl_pct / 100) ** (365.0 / days) - 1) * 100, 2)
    except (ValueError, OverflowError):
        return None


def build_portfolio_snapshot(name: str) -> list[dict]:
    df = load_portfolio(name)
    if df.empty:
        return []
    tickers = df["Ticker"].unique().tolist()
    prices = fetch_current_prices(tickers)
    rows = []
    for _, row in df.iterrows():
        t = row["Ticker"]
        shares = float(row["Shares"])
        cb = float(row["Cost Basis"])
        purchase_date = str(row.get("Purchase Date", "") or "").strip()
        days = _days_held(purchase_date)
        price = prices.get(t)
        if price is None:
            mv = pnl_d = pnl_p = ann_r = None
        else:
            mv = round(shares * price, 2)
            cost = shares * cb
            pnl_d = round(mv - cost, 2)
            pnl_p = round(pnl_d / cost * 100, 2) if cost else None
            ann_r = _annualized_return(pnl_p, days)
        rows.append({
            "id": str(row["ID"]),
            "ticker": t,
            "shares": shares,
            "cost_basis": cb,
            "purchase_date": purchase_date or None,
            "days_held": days,
            "current_price": round(price, 2) if price else None,
            "market_value": mv,
            "pnl_dollars": pnl_d,
            "pnl_pct": pnl_p,
            "ann_return_pct": ann_r,
        })
    return rows


def build_portfolios_summary() -> list[dict]:
    names = list_portfolio_names()
    if not names:
        return []
    portfolios = {n: load_portfolio(n) for n in names}
    all_tickers = list({t for df in portfolios.values() for t in df["Ticker"].tolist()})
    prices = fetch_current_prices(all_tickers) if all_tickers else {}
    returns = fetch_one_year_returns(all_tickers) if all_tickers else {}
    summaries = []
    for name, df in portfolios.items():
        total_value = total_cost = weighted_return = 0.0
        for _, row in df.iterrows():
            t = row["Ticker"]
            shares, cb = float(row["Shares"]), float(row["Cost Basis"])
            total_cost += shares * cb
            price = prices.get(t)
            if price:
                mv = shares * price
                total_value += mv
                weighted_return += mv * (returns.get(t) or 0)
        pnl = total_value - total_cost
        summaries.append({
            "name": name,
            "total_value": round(total_value, 2),
            "total_cost": round(total_cost, 2),
            "pnl_dollars": round(pnl, 2),
            "pnl_pct": round(pnl / total_cost * 100, 2) if total_cost else None,
            "return_1y_pct": round(weighted_return / total_value, 2) if total_value else None,
        })
    return summaries


# ── Excel import / export ────────────────────────────────────────────────────

_TEMPLATE_COLS = ["Ticker", "Shares", "Cost Basis", "Purchase Date", "Portfolio"]
_EXAMPLES = [
    ("AAPL",  10,   150.00, "2023-01-15", "Main"),
    ("MSFT",   5,   280.00, "2022-08-20", "Main"),
    ("NVDA",   8,   220.00, "",           "Tech"),
    ("GOOGL",  3,   120.00, "2024-03-10", "Tech"),
]


def generate_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"

    hdr_fill = PatternFill(fgColor="1E2332", fill_type="solid")
    hdr_font = Font(bold=True, color="E2E8F0", size=11)
    ex_fill  = PatternFill(fgColor="1A1F2E", fill_type="solid")
    ex_font  = Font(color="64748B", italic=True, size=10)
    note_font = Font(color="475569", italic=True, size=9)

    for col, name in enumerate(_TEMPLATE_COLS, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r, row in enumerate(_EXAMPLES, 2):
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = ex_fill
            c.font = ex_font

    note_row = len(_EXAMPLES) + 3
    note = ws.cell(row=note_row, column=1,
                   value="* Rows 2–5 are examples — replace or delete them.  "
                         "Purchase Date is optional; use YYYY-MM-DD format.  "
                         "Portfolio names are created automatically if they don't exist.")
    note.font = note_font
    ws.merge_cells(start_row=note_row, start_column=1,
                   end_row=note_row, end_column=len(_TEMPLATE_COLS))

    col_widths = [12, 10, 14, 18, 16]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_date(raw) -> str:
    """Return YYYY-MM-DD string or '' if raw is empty/unparseable."""
    if raw is None:
        return ""
    if hasattr(raw, "strftime"):          # datetime/date from openpyxl
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if not s:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""                             # unparseable — import without date


def import_from_excel(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "skipped": 0, "errors": []}

    # Normalise header names
    raw_headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    def col(name):
        for alias in (name, name.replace(" ", "_"), name.replace("_", " ")):
            try:
                return raw_headers.index(alias)
            except ValueError:
                pass
        return None

    ci = {
        "ticker":        col("ticker"),
        "shares":        col("shares"),
        "cost_basis":    col("cost basis") if col("cost basis") is not None else col("cost_basis"),
        "purchase_date": col("purchase date") if col("purchase date") is not None else col("purchase_date"),
        "portfolio":     col("portfolio"),
    }
    missing = [k for k, v in ci.items() if v is None and k not in ("purchase_date",)]
    if missing:
        raise ValueError(f"Template is missing required columns: {', '.join(missing)}")

    existing_portfolios = set(list_portfolio_names())
    imported = skipped = 0
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        def get(key):
            idx = ci.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        ticker    = str(get("ticker") or "").strip().upper()
        portfolio = str(get("portfolio") or "").strip()

        if not ticker or not portfolio:
            skipped += 1
            continue

        # Skip the built-in example rows if the user left them in
        if ticker in {t for t, *_ in _EXAMPLES} and portfolio in {p for *_, p in _EXAMPLES}:
            skipped += 1
            continue

        try:
            shares     = float(get("shares"))
            cost_basis = float(get("cost_basis"))
        except (TypeError, ValueError):
            errors.append(f"Row {row_num} ({ticker}): Shares and Cost Basis must be numbers.")
            continue

        if shares <= 0 or cost_basis <= 0:
            errors.append(f"Row {row_num} ({ticker}): Shares and Cost Basis must be positive.")
            continue

        purchase_date = _parse_date(get("purchase_date"))

        if portfolio not in existing_portfolios:
            create_portfolio(portfolio)
            existing_portfolios.add(portfolio)

        try:
            add_holding(portfolio, ticker, shares, cost_basis, purchase_date)
            imported += 1
        except Exception as e:
            errors.append(f"Row {row_num} ({ticker}): {e}")

    return {"imported": imported, "skipped": skipped, "errors": errors}
